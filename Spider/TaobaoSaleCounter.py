#! python3
# encoding: utf-8

try:
    import os
    import json
    import requests
    import re
    import bs4
    import pandas as pd
    import progressbar
    import openpyxl
    import logging
    from io import BytesIO
    from PIL import Image as Im
    from openpyxl.drawing.image import Image
except Exception as error:
    print("MISSING SOME MODULE(s)")
    print (error)
    os.system("pip install pandas progressbar2 requests openpyxl pillow ")
    print("TRY TO INSTALL SOME MODs")
    print("PLEASE UPGRADE PIP IF IT DOESN'T WORK ")
    print("Restart this Program!")
    exit(-2) 

logging.basicConfig(filename='log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# Get comment url
def getUrl(currentPageNum):
    url = 'https://rate.taobao.com/feedRateList.htm?auctionNumId=%s&userNumId=%s&currentPageNum=%s' % (auctionNumId, userNumId, currentPageNum)
    return url

# Get web data
def getHtml(url):
    headers = {
        'Connection': 'Keep-Alive',
        'Accept': 'text/html, application/xhtml+xml, */*',
        'Accept-Language': 'en-US,en;q=0.8,zh-Hans-CN;q=0.5,zh-Hans;q=0.3',
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; rv:11.0) like Gecko'
    }
    res = requests.get(url, headers = headers, timeout = 10)
    res.raise_for_status()	
    return res

itemUrl = input('Enter Taobao Item Url - ') # Taobao Item Url
auctionNumId = re.findall('id=([0-9]+)', itemUrl)[0] # Item ID
res = getHtml(itemUrl)
userNumId = re.findall('userid=([0-9]+)', res.text)[0]	# Sellers' ID

# MaxPage: 250
pageNum = input('Input page number: ')
if len(pageNum) < 1 or int(pageNum) > 250:
    pageNum = 250
    print('Sorry! I can only get 250 pages!')
urls = [getUrl(currentPageNum) for currentPageNum in range(1, int(pageNum)+1)]

# Get sku and image url
skuImg = {} 
soup = bs4.BeautifulSoup(res.text, "html.parser")
cels = soup.select('li[data-value] > a')
for cel in cels:
    img_data = cel.get('style')
    if img_data == None:
        logging.debug(cel.getText().strip('\n'))
        continue
    c_url = 'https:' + re.findall('(//.*?\.[a-z]{2}g?)_\d\dx\d\d', img_data)[0]
    skuImg[cel.getText().strip('\n')] = c_url

# Get sku and the number of buyers
commentsNum = 0
nullNum = 0
saleCounter = {}
bar = progressbar.ProgressBar()

for n in bar(range(len(urls))):
    req = getHtml(urls[n])
    if len(req.text) < 500:
        print('\nmaxPage:%s' % n)
        break
    myweb = json.loads(req.text[3:len(req.text)-2])
    comments = myweb['comments']
    commentsNum += len(comments)
    for i in range(len(comments)):
        try:
            sku = comments[i]['auction']['sku'].split(':')[1]
            if '[' in sku:
                sku = sku.split('[')[0]
        except:
            nullNum += 1
            logging.debug("Sku is Null! \nOn page: %s\nLocation: %s" % (urls[n], i)) # Sku is null!
            continue
        saleCounter[sku] = saleCounter.get(sku,0) + 1

#  Writing to MS Excel	
df = pd.DataFrame({'counter': pd.Series(saleCounter, index=list(saleCounter.keys()))})
df.sort_values(by='counter', ascending=False, inplace=True)
fname = 'saleCounter_%s.xlsx' % auctionNumId
df.to_excel(fname, sheet_name='Sheet1')

# Append sku picture to Ms Excel
wb = openpyxl.load_workbook(fname)
sheet = wb.active
pbar = progressbar.ProgressBar()
for rowNum in pbar(range(2, sheet.max_row+1)):
    sheet.row_dimensions[rowNum].height = 180
    sheet.column_dimensions['C'].width = 30
    try:
        imgUrl = skuImg[sheet.cell(row=rowNum, column=1).value]
    except:
        logging.debug(sheet.cell(row=rowNum, column=1).value)
        continue
    reqs = getHtml(imgUrl)
    data = reqs.content
    im = Im.open(BytesIO(data))
    resizeImg = im.resize((230, 230))
    os.makedirs('tmp', exist_ok=True)
    fileName = 'C%s' % rowNum + '.' + imgUrl.split('.')[-1]
    file =  os.path.join('tmp', fileName)
    resizeImg.save(file)
    img = Image(file)
    sheet.add_image(img, 'C%s' % rowNum)

wb.save(fname)
print('Total comments: %s || Null comments: %s' % (commentsNum, nullNum))
print("\nHi~Finished~~~")
